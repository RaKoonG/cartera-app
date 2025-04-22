import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from win32com.client import Dispatch
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import streamlit as st
from datetime import datetime
import tempfile
import uuid

# Función para convertir xls a xlsx
def convertir_xls_a_xlsx(archivo):
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    libro = excel.Workbooks.Open(archivo)
    nuevo_archivo = archivo + "x"
    libro.SaveAs(nuevo_archivo, FileFormat=51)
    libro.Close(False)
    excel.Quit()
    return nuevo_archivo

# Procesar archivo
def procesar_archivo(ruta_archivo):
    if ruta_archivo.endswith(".xls"):
        ruta_archivo = convertir_xls_a_xlsx(ruta_archivo)

    df = pd.read_excel(ruta_archivo)
    df = df.iloc[7:]

    columnas_a_eliminar = [1, 3, 4, 5, 6]
    df.drop(df.columns[columnas_a_eliminar], axis=1, inplace=True)
    df.columns = df.columns[:].to_list()
    df.rename(columns={df.columns[1]: "Cliente"}, inplace=True)

    df = df[df["No. Venta"].notna()]
    df = df[df["Cliente"].notna()]
    df = df[df["Vence"].notna()]
    df = df[~df["Clasificacion"].isin(["ABOGADO", "DCL-PRELEGAL", "EMPLEADOS", "PRE-LEGAL"])]
    df = df[~df["Cliente"].str.contains("CUENTAS INCOBRABLES|INCOBRABLES", na=False)]

    filtrados = df[df["Clasificacion"].str.contains("C.IMPULSA|F. TAMAZULA|FINANCIERA", na=False)]
    df = df[~df.index.isin(filtrados.index)]

    def insertar_filas_vacias(dataframe):
        df_ordenado = dataframe.sort_values(by="No. Cliente").reset_index(drop=True)
        nuevas_filas = []
        for i in range(len(df_ordenado)):
            nuevas_filas.append(df_ordenado.iloc[i])
            if i < len(df_ordenado) - 1:
                if df_ordenado.iloc[i]["No. Cliente"] != df_ordenado.iloc[i + 1]["No. Cliente"]:
                    nuevas_filas.append(pd.Series([None] * len(df_ordenado.columns), index=df_ordenado.columns))
        return pd.DataFrame(nuevas_filas)

    df_con_filas = insertar_filas_vacias(df)
    filtrados_con_filas = insertar_filas_vacias(filtrados)

    # Colorear fechas
    def formato_color(fecha):
        if pd.isna(fecha):
            return "black"
        hoy = datetime.now()
        if fecha.month < hoy.month or fecha.year < hoy.year:
            return "red"
        elif fecha.month == hoy.month and fecha.year == hoy.year:
            return "black"
        else:
            return "green"

    def agregar_estatus(df):
        df["Estatus"] = ""
        df["Grupo"] = df["No. Cliente"].fillna(method="ffill")  # para agrupar con filas vacías

        for grupo, datos in df.groupby("Grupo"):
            fechas = datos["Vence"].dropna()
            hoy = datetime.now()
            moroso = fechas[fechas.apply(lambda f: f.month < hoy.month or f.year < hoy.year)]
            no_moroso = fechas[fechas.apply(lambda f: f.month > hoy.month or f.year > hoy.year)]
            del_mes = fechas[fechas.apply(lambda f: f.month == hoy.month and f.year == hoy.year)]

            if not moroso.empty:
                estatus = "MOROSO"
            elif not moroso.empty and not no_moroso.empty:
                estatus = "MOROSO"
            elif no_moroso.empty and del_mes.empty:
                estatus = ""
            elif not no_moroso.empty and del_mes.empty:
                estatus = "NO MOROSO"
            else:
                estatus = "DOS"

            df.loc[df["Grupo"] == grupo, "Estatus"] = estatus

        df.drop(columns=["Grupo"], inplace=True)
        return df

    df_final = agregar_estatus(df_con_filas)
    filtrados_final = agregar_estatus(filtrados_con_filas)

    # Guardar archivo temporal
    temp_dir = tempfile.gettempdir()
    nombre_archivo = f"cartera_procesada_{uuid.uuid4().hex[:6]}.xlsx"
    ruta_guardado = os.path.join(temp_dir, nombre_archivo)

    with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name="Hoja1", index=False)
        filtrados_final.to_excel(writer, sheet_name="Filtrados", index=False)

        for hoja in writer.sheets.values():
            for col_idx, col in enumerate(hoja.iter_cols(min_row=2), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                hoja.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            for row in hoja.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter == 'C':  # Columna Vence
                        color = formato_color(cell.value)
                        if color == "red":
                            cell.font = Font(color="FF0000")
                        elif color == "green":
                            cell.font = Font(color="008000")
                        else:
                            cell.font = Font(color="000000")

    return ruta_guardado

# ----------------- INTERFAZ STREAMLIT -----------------
st.title("Procesador de Cartera")
archivo = st.file_uploader("Sube tu archivo .xls o .xlsx", type=["xls", "xlsx"])

if archivo is not None:
    ruta_temporal = os.path.join(tempfile.gettempdir(), archivo.name)
    with open(ruta_temporal, "wb") as f:
        f.write(archivo.read())

    st.write("Procesando archivo...")
    try:
        ruta_resultado = procesar_archivo(ruta_temporal)
        with open(ruta_resultado, "rb") as f:
            st.download_button("Descargar archivo procesado", f, file_name="cartera_procesada.xlsx")
        st.success("¡Archivo procesado exitosamente!")
    except Exception as e:
        st.error(f"Error durante el procesamiento: {e}")
